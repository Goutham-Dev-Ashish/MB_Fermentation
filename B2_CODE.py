import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

# Specify the path to your Excel file
file_path = '/Users/goutham/Desktop/Mercury Bio/Fermentation Data/MB24/MB_24_Power Query/MB24 B2/MB24-B2/MB24-26-B2/MB24-26-B2_PLOTTED.xlsx'

# Load the workbook and select the active sheet
workbook = load_workbook(file_path)
worksheet = workbook.active

# Function to create and add scatter chart
def add_scatter_chart(sheet, chart_title, x_axis_title, y_axis_title, x_col, y_cols, y_titles, colors, anchor_row, anchor_col):
    chart = ScatterChart()
    chart.title = chart_title
    chart.style = 13
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.majorUnit = 5

    x_values = Reference(sheet, min_col=x_col, min_row=2, max_row=sheet.max_row)

    for y_col, title, color in zip(y_cols, y_titles, colors):
        y_values = Reference(sheet, min_col=y_col, min_row=2, max_row=sheet.max_row)
        series = Series(y_values, x_values, title=title)
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 20000  # Width in EMUs (1 point = 12700 EMUs)
        chart.series.append(series)

    chart.anchor = f"{anchor_col}{anchor_row}"  # Set the anchor position
    sheet.add_chart(chart)

run_number = "MB24-26_B2"

# Add all charts to the original sheet
charts_info = [
    (worksheet, run_number + " E-CO2, E-O2", 'Process Time', ' ', 1, [4, 5], ["ECO2", "EO2"], ["0000FF", "FF0000"], worksheet.max_row + 2, 'A'),
    (worksheet, run_number + " DO, LPM", 'Process Time', ' ', 1, [10, 2], ["pO2", "AIRSPEED"], ["0000FF", "FF0000"], worksheet.max_row + 2, 'H'),
    (worksheet, run_number + " DO,rpm", 'Process Time', ' ', 1, [10, 11], ["pO2", "STIRRER"], ["0000FF", "FF0000"], worksheet.max_row + 2, 'O'),
    (worksheet, run_number + " pH,Base", 'Process Time', ' ', 1, [8, 3], ["pH", "Base"], ["0000FF", "FF0000"], worksheet.max_row + 2, 'V'),
    (worksheet, run_number + " GLC_FEEDS", 'Process Time', ' ', 1, [7, 13, 15], ["F_WEIGHT", "SUBS_B1", "VWEIGHT"], ["0000FF", "FF0000", "00FF00"], worksheet.max_row + 20, 'A'),
    (worksheet, run_number + " TEMP", 'Process Time', ' ', 1, [14], ["TEMP"], ["0000FF"], worksheet.max_row + 20, 'H'),
    (worksheet, run_number + " DO:-SAT,VAL", 'Process Time', ' ', 1, [10, 9], ["pO2_%SAT", "pO2_VAL"], ["0000FF", "FF0000"], worksheet.max_row + 20, 'O'),
    (worksheet, run_number + " FEED:Rate,mL", 'Process Time', ' ', 1, [12, 13], ["Feed Rate", "Feed mL"], ["0000FF", "FF0000"], worksheet.max_row + 20, 'V')
]

for chart_info in charts_info:
    add_scatter_chart(*chart_info)

# Save the workbook once after adding all charts
workbook.save(file_path)

# Create a new blank sheet and name it
new_sheet = workbook.create_sheet(title="RQ")

# Copy columns 1, 4, and 5 to the new sheet
for row in range(1, worksheet.max_row + 1):
    new_sheet.cell(row=row, column=1).value = worksheet.cell(row=row, column=1).value  # Column 1 (Process Time)
    new_sheet.cell(row=row, column=2).value = worksheet.cell(row=row, column=4).value  # Column 4 (E-CO2)
    new_sheet.cell(row=row, column=3).value = worksheet.cell(row=row, column=5).value  # Column 5 (E-O2)

# Add label to cell H1 and add the MIN formula
new_sheet['H1'] = 'CO2 MIN'
new_sheet['H2'] = f"=MIN(B2:B{worksheet.max_row})"

# Add label to cell I1 and add the MAX formula
new_sheet['I1'] = 'O2 MAX'
new_sheet['I2'] = f"=MAX(C2:C{worksheet.max_row})"

# Label cell D1 as CER and add the formula for D2 onwards
new_sheet['D1'] = 'CER'
for row in range(2, worksheet.max_row + 1):
    new_sheet[f'D{row}'] = f"=B{row}-H$2"

# Label cell E1 as OUR and add the formula for E2 onwards
new_sheet['E1'] = 'OUR'
for row in range(2, worksheet.max_row + 1):
    new_sheet[f'E{row}'] = f"=I$2-C{row}"

# Label cell F1 as RQ and add the formula for F2 onwards
new_sheet['F1'] = 'RQ'
for row in range(2, worksheet.max_row + 1):
    new_sheet[f'F{row}'] = f"=D{row}/E{row}"

# Add scatter chart to the new sheet
add_scatter_chart(
    new_sheet,
    chart_title=run_number + " RQ",
    x_axis_title='Process Time',
    y_axis_title=' ',
    x_col=1,
    y_cols=[4, 5, 6],  # Adjusted to match the new sheet's columns
    y_titles=["CER", "OUR", "RQ"],
    colors=["0000FF", "FF0000", "00FF00"],
    anchor_row=new_sheet.max_row + 2,
    anchor_col='A'
)

# Add another scatter chart with Process Time on X-axis and OUR on Y-axis
def add_our_chart(sheet, chart_title, x_axis_title, y_axis_title, x_col, y_col, y_titles, anchor_row, anchor_col):
    chart = ScatterChart()
    chart.title = chart_title
    chart.style = 13
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.majorUnit = 5

    x_values = Reference(sheet, min_col=x_col, min_row=2, max_row=sheet.max_row)
    y_values = Reference(sheet, min_col=y_col, min_row=2, max_row=sheet.max_row)
    series = Series(y_values, x_values, title="RQ")
    series.graphicalProperties.line.solidFill = "00FF00"  # Green color
    series.graphicalProperties.line.width = 20000  # Width in EMUs (1 point = 12700 EMUs)
    chart.series.append(series)

    chart.anchor = f"{anchor_col}{anchor_row}"  # Convert column index to letter
    sheet.add_chart(chart)

# Add OUR chart to the new sheet
add_our_chart(
    new_sheet,
    chart_title=run_number + " Process Time vs OUR",
    x_axis_title='Process Time',
    y_axis_title='OUR',
    x_col=1,
    y_col=5,
    y_titles=["OUR"],
    anchor_row=new_sheet.max_row + 20,
    anchor_col='A'
)

# Save the workbook after making all changes
workbook.save(file_path)

print("Scatter plots with straight lines have been added to the Excel file after the last row of data.")
